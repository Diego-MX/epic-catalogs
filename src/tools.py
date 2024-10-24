"""Low-level tools to be used inside the app. 
"""
# pylint: disable=missing-class-docstring
# pylint: disable=too-few-public-methods
# pylint: disable=raising-bad-type

from base64 import b64encode as enc64
import re
import os
from time import time

from openpyxl import load_workbook, utils as xl_utils 
import pandas as pd
from sqlalchemy import create_engine
from unidecode import unidecode


def dict_get(a_dict: dict, keys_ls, val_else=None): 
    for a_key in keys_ls: 
        if a_key in a_dict: 
            a_val = a_dict[a_key] 
            break
    else: 
        if val_else: 
            a_val = val_else
        else: 
            raise ValueError('Keys not found, and value not provided.')
    return a_val


def str_camel_to_snake(cameled: str):
    subbed = re.sub('(.)([A-Z][a-z]+)',  r'\1_\2', cameled)
    snaked = re.sub('([a-z0-9])([A-Z])', r'\1_\2', subbed).lower()
    return snaked


def str_snake_to_camel(snaked: str, first_word_too=False, decode=False):
    if decode:
        snaked = unidecode(snaked)
    splitted = snaked.split('_')
    first_word = splitted.pop(0)
    first_camel = first_word.title() if first_word_too else first_word.lower()
    cameled = first_camel + ''.join(word.title() for word in splitted)
    return cameled


def shortcut_target(filename, file_ext=None):
    def ext_regex(file_ext):
        if file_ext is None: 
            file_ext = 'xlsx'
        if isinstance(file_ext, str):
            ext_reg = file_ext
        elif isinstance(file_ext, list):
            ext_reg = f"{'|'.join(file_ext)}"
        else:
            raise ValueError('FILE_EXT format is not supported.')
        return ext_reg
    
    file_regex = fr'C:\\.*\.{ ext_regex(file_ext) }'
    with open(filename, 'r', encoding='ISO-8859-1') as _f: 
        a_path = a_path = re.findall(file_regex, _f.read(), flags=re.DOTALL)

    if len(a_path) != 1: 
        raise 'Not unique or No shortcut targets found in link.'    
    return a_path[0]


def read_excel_table(file, sheet, table): 
    try:
        a_wb = load_workbook(file, data_only=True)
    except xl_utils.exceptions.InvalidFileException: 
        a_wb = load_workbook(shortcut_target(file), data_only=True)
    a_ws  = a_wb[sheet]
    a_tbl = a_ws.tables[table]
    
    rows_ls = [[cell.value for cell in row] for row in a_ws[a_tbl.ref]]
    tbl_df  = pd.DataFrame(data=rows_ls[1:], index=None, columns=rows_ls[0])
    return tbl_df



def dict_minus(a_dict, key, copy=True): 
    b_dict = a_dict.copy() if copy else a_dict
    b_dict.pop(key)
    return b_dict


def encode64(a_str): 
    encoded = enc64(a_str.encode('ascii')).decode('ascii')
    return encoded


def dataframe_response(a_df, cols_df=None, resp_keys=None, drop_nas=True): 
    # Convierte A_DF en {numberOfRecords, attributes, recordSet, pagination}
    # COLS_DF tiene columnas:
    #  DTIPO['datetime', 'date', ...]
    #  ATRIBUTO
    if resp_keys is None:
        resp_keys = {}

    if cols_df is None:
        the_types = {'bool': 'logical', 'object': 'character'}
        cols_df = (a_df.dtypes.replace(the_types)
            .to_frame('dtipo').reset_index()
            .rename(columns={'index': 'nombre'}))
    
    ts_cols     = cols_df.query('dtipo == "datetime"')['nombre']
    date_cols_0 = cols_df.query('dtipo == "date"')['nombre']

    date_assign = {col: a_df[col].apply(lambda dt: dt.strftime('%Y-%m-%d')) 
            for col in date_cols_0 if col in a_df.columns.values}
    str_assign  = {col: str 
            for col in ts_cols if col in a_df.columns.values}
    
    b_df  = a_df.astype(str_assign).assign(**date_assign)
    if drop_nas:
        b_records = [row.dropna().to_dict() for _i, row in b_df.iterrows()]
    else: 
        b_records = b_df.fillna('').to_dict(orient='records')

    std_dict = {
        'numberOfRecords' : b_df.shape[0],
        'attributes'      : b_df.columns.tolist(),
        'recordSet'       : b_records,
        'pagination'      : {'hasPagination': False}}

    df_response = {resp_keys.get(kk, kk): vv for kk,vv in std_dict.items()}
    return df_response


def type_environment():
    """
    Función que muestra sí el programa se encuentra en el ambiente Docker o en Local.
    """
    environment = (os.path.exists('/.dockerenv') or os.path.isfile('/proc/1/cgroup') and
                   'docker' in open('/proc/1/cgroup', encoding="utf-8").read())
    return environment


def get_connection():
    """
    Conexión con la DB en Azure, se crea una bifurcasión dado que 
    se trabaja en docker y de forma local"""

    params = {
        'username' : "USUARIO_BATALLA",
        'password' : "18.07.2024.JUIK.juik", 
        'server' : "sql-lakehylia-dev.database.windows.net", 
        'database' : "webapp_catalogs"}

    if type_environment():
        params['el_driver'] = "/opt/microsoft/msodbcsql18/lib64/libmsodbcsql-18.4.so.1.1"
    else:
        params['el_driver'] = "ODBC+Driver+18+for+SQL+Server"
    connection_string = 'mssql+pyodbc://{username}:{password}@{server}/{database}?driver={el_driver}'.format(**params)

    if connection_string is None:
        raise ValueError("La variable de entorno CONNECTION_STRING no está configurada.")

    connection = create_engine(connection_string)

    return connection


class Timer:
    """Simple class to print timed processes."""
    def __init__(self, print_mode=0): 
        """print_mode: 
            0 : Dont print
            1 : since last call
            -1: since beginning
        """
        self.timers = [time()]
        self.print_mode = print_mode
    
    def set_mark(self, desc, and_print=0): 
        self.timers.append(time())

        and_print = and_print or self.print_mode
        if and_print == 0: 
            return
        
        if and_print == 1:  # Doesnt reference indices, but timeframe. 
            ellapsed = self.timers[-1] - self.timers[-2]    
        elif and_print == -1:
            ellapsed = self.timers[-1] - self.timers[0]
        else: 
            raise ValueError("'and_print/print_mode' can be one of [0, 1, -1]")
        print(f"{desc}:\t{ellapsed:.2f} seconds")
